__author__ = 'Vishal'

import os
import sys
import click
import MySQLdb as sqlDB
import openpyxl as xl
import django
from OnlineProject import settings
os.environ['DJANGO_SETTINGS_MODULE'] = 'OnlineProject.settings'
django.setup()
from OnlineApp.models import *


@click.group()
def onlinedb():
    """Commands for db import/export"""
    pass


def connect():
    """Connects to MySQL server."""
    try:
        db = sqlDB.connect(host="localhost", user="root", passwd="dbpass")
        return db
    except sqlDB.Error:
        return None


def connect_db(dbname):
    """Connects to the given database of MySQL server."""
    try:
        db = sqlDB.connect(host="localhost", user="root", passwd="dbpass", db=dbname)
        return db
    except sqlDB.Error:
        return None


def get_name(name):
    name = name[:name.find("_mock")]
    name = name[len(name) - name[::-1].find("_"):]
    return name


def load_college_data(colleges_ws):
    for rows in range(2, colleges_ws.max_row + 1):
        data = [[]] * colleges_ws.max_column
        for cols in range(1, colleges_ws.max_column + 1):
            data[cols - 1] = colleges_ws.cell(row=rows, column=cols).value
        c = College(name=data[0], location=data[2], acronym=data[1], contact=data[3])
        c.save()


def load_students_data(student_ws):
    for rows in range(2, student_ws.max_row + 1):
        data = [[]] * student_ws.max_column
        for cols in range(1, student_ws.max_column + 1):
            data[cols-1] = student_ws.cell(row=rows, column=cols).value
        c = College.objects.get(acronym=data[1])
        s = Student(name=data[0], email=data[2], db_folder=data[3], college=c)
        s.save()


def load_dropped_students_data(dropped_student_ws):
    for rows in range(2, dropped_student_ws.max_row + 1):
        data = [[]] * dropped_student_ws.max_column
        for cols in range(1, dropped_student_ws.max_column + 1):
            data[cols-1] = dropped_student_ws.cell(row=rows, column=cols).value
        c = College.objects.get(acronym=data[1])
        s = Student(name=data[0], email=data[2], dropped_out=True, db_folder=data[3], college=c)
        s.save()


def load_marks_data(marks_ws):
    for rows in range(2, marks_ws.max_row+1):
        data = [[]] * marks_ws.max_column
        for cols in range(1, marks_ws.max_column+1):
            data[cols-1] = marks_ws.cell(row=rows, column=cols).value
            if cols == 1:
                data[0] = get_name(data[0].lower())
        s = Student.objects.get(db_folder=data[0])
        m = MockTest1(problem1=data[1], problem2=data[2], problem3=data[3], problem4=data[4],
                      total=data[5], student=s)
        m.save()


@onlinedb.command(name='createdb', short_help="creates the db in django settings using the credentials specified there")
def createdb():
    db = connect()
    dbname = settings.DATABASES['default']['NAME']
    if db is None:
        click.echo("Error occurred")
        sys.exit(1)

    click.echo("Creating a database ...")

    sql_query = db.cursor()
    sql_cmd = f'CREATE DATABASE IF NOT EXISTS {dbname}'
    sql_query.execute(sql_cmd)
    db.commit()
    db.close()

    click.echo("Database is created with the name %s" % dbname)
    click.echo("Creating a table ...")

    os.system('python manage.py makemigrations')
    os.system('python manage.py migrate')

    click.echo('Tables are created ...')


@onlinedb.command(name='dropdb', short_help="drops the db in django settings using the credentials specified there")
def dropdb():
    """Drops the database."""

    db = connect()
    dbname = settings.DATABASES['default']['NAME']

    if db is None:
        click.echo("Error occurred")
        sys.exit(1)

    click.echo("Dropping a database ...")

    sql_query = db.cursor()
    sql_cmd = f'DROP DATABASE IF EXISTS {dbname}'
    sql_query.execute(sql_cmd)
    db.commit()
    db.close()

    click.echo("Database %s is dropped" % dbname)


@onlinedb.command(name='cleardata', short_help="clears the data from the database")
def clear_data():
    """Clears the data from the database by truncating all the tables."""

    click.echo('Deleting data from tables...')
    College.objects.all().delete()
    click.echo('Data deleted...')


@onlinedb.command(name='populate', short_help='Fills the data into the database')
def populate():
    student_wb = xl.load_workbook('C://PythonCourse//students.xlsx')
    marks_wb = xl.load_workbook('C://PythonCourse//marks.xlsx')

    click.echo('Finished loading source files...')
    click.echo('Loading colleges data...')

    colleges_ws = student_wb['Colleges']
    load_college_data(colleges_ws)

    click.echo('Colleges data loaded...')
    click.echo('Loading students data...')

    student_ws = student_wb['Current']
    load_students_data(student_ws)

    click.echo('Finished loading students data...')
    click.echo('Loading dropped students data...')

    dropped_student_ws = student_wb['Deletions']
    load_dropped_students_data(dropped_student_ws)

    click.echo('Finished loading dropped students data...')
    click.echo('Loading marks data...')

    marks_ws = marks_wb['Sheet']
    load_marks_data(marks_ws)

    click.echo('Finished loading marks data...')


if __name__ == '__main__':
    onlinedb()
